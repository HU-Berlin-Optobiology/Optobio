# -*- coding: utf-8 -*-
"""
Created on Fri Feb 28 09:43:41 2025

@author: Erich
"""

from tkinter import ttk
from tkinter import filedialog, messagebox
from tkinter import *
import pandas as pd
import numpy as np
import openpyxl
import threading

# Initiating main window and everything else
root = Tk()
root.geometry('630x400')
root.config(bg="black")
root.focus_force()
root.title("ExcelVergleich")
####################################################################################################################################
# Styles

style = ttk.Style()
style.configure("C.TLabel", font=('Arial', 10), foreground="white", background="black")
style.configure("C.TEntry", font=("Arial", 10))
style.configure("CExit.TButton", font=("Arial", 10, "underline"), )
style.configure("C.TButton", font=("Arial", 10))
style.configure("C.TFrame", background="black")

####################################################################################################################################
# Params
active0 = False
active1 = False
active2 = False
start0 = False
start1 = False
start2 = False
file_type = [("Excel", "*.xlsx")]

####################################################################################################################################
# Functions

def destroyRoot(*event):
    root.destroy()

def loadExcelUser(*event):
    global  LoadUser
    global active0
    LoadUser = filedialog.askopenfilename(parent=root, title="Die Excel-Tabelle in der HU-Box auswählen", initialdir="Z:/HU_Box/orga lists (AB, constructs)", filetypes=file_type, initialfile="")
    path_load_user.set(LoadUser)
    active0 = True
    start0 = True
    if active0 and active1 and start0:
        threading.Thread(target=LoadTables, daemon=True).start()
    if active0 and active1 and active2:
        CompareBtn.config(state=ACTIVE)
    
def loadExcelBackup(*event):
    global  LoadBackup
    global active1
    LoadBackup = filedialog.askopenfilename(parent=root, title="Die originale Excel-Tabelle auswählen", initialdir="C:/Users/", filetypes=file_type)
    path_load_backup.set(LoadBackup)
    active1 = True
    start1 = True
    if active0 and active1 and start1:
        threading.Thread(target=LoadTables, daemon=True).start()
    if active0 and active1 and active2:
        CompareBtn.config(state=ACTIVE)
    
def saveResult(*event):
    global Save
    global active2
    InitDir = ("/").join(LoadBackup.split("\\")[:-1])
    Save = filedialog.asksaveasfilename(parent=root, title="Wo die resultierende Excel-tabelle gespeichert werden soll und unter welchem Namen", initialdir=InitDir, defaultextension=".xlsx",
                                        filetypes=file_type)
    path_save.set(Save)
    active2 = True
    if active0 and active1 and active2:
        CompareBtn.config(state=ACTIVE)

def ShowWork(*event):
    global thread1
    p_bar.grid(row=4, column=0, padx=10, pady=10, sticky=W, columnspan=4)
    p_bar.start()
    thread1 = threading.Thread(target=compareExcel, daemon=True)
    thread1.start()
    
def LoadTables():
    global xl1
    global xl2
    global chosen
    file1 = LoadBackup
    xl1 = pd.ExcelFile(file1)
    file2 = LoadUser
    xl2 = pd.ExcelFile(file2)
    sheets1 = xl1.sheet_names
    sheets2 = xl2.sheet_names
    chosen = StringVar()
    if sheets1!=sheets2:
        messagebox.showerror("Nicht identische Tabellen", "Die Tabellen haben unterschiedliche Benennungen der Arbeitsblätter.")
        active0 = False
        active1 = False
        start0 = False
        start1 = False
    else:
        # Drop Down menu for sheets
        sheets = ["Arbeitsblatt"]+sheets1+["Alle"]
        SheetsDrop = ttk.OptionMenu(root, chosen, *sheets)
        SheetsDrop.grid(row=3, column=1, columnspan=3, padx=10, pady=10, sticky=W)
    
def compareExcel():
    global wb
    compare = chosen.get()
    if compare == "Alle":
        sheets = xl1.sheet_names
        for sheet in sheets:
            df1 = pd.read_excel(xl1, sheet_name=sheet).replace(np.nan, "\xa0")
            df2 = pd.read_excel(xl2, sheet_name=sheet).replace(np.nan, "\xa0")
            if sheet == sheets[0]:
                #init of excel sheet and book
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet+"_Fehlend"
                keys = df1.keys()
                r=1
                c=1
                for i in range(len(keys)):
                    value = keys[i]
                    ws.cell(row=r, column=c+i, value=value)
                r+=1
                #end of init
            elif sheet != sheets[0]:
                #init of excel sheet and book
                ws = wb.create_sheet(sheet+"Fehlend")
                keys = df1.keys()
                r=1
                c=1
                for i in range(len(keys)):
                    value = keys[i]
                    ws.cell(row=r, column=c+i, value=value)
                r+=1
                #end of init
            for i in range(len(df1)):
                for j in range(len(df2)):
                    backup = df1.loc[i].tolist()
                    user = df2.loc[j].tolist()
                    if backup==user:
                        break
                    elif j==len(df2)-1:
                        for k in range(len(backup)):
                            value=backup[k]
                            ws.cell(row=r, column=c+k, value=value)
                        r+=1
    elif compare != "Alle":  
        df1 = pd.read_excel(xl1, sheet_name=compare).replace(np.nan, "\xa0")
        df2 = pd.read_excel(xl2, sheet_name=compare).replace(np.nan, "\xa0")
        #init of excel sheet and book
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fehlend"
        keys = df1.keys()
        r=1
        c=1
        for i in range(len(keys)):
            value = keys[i]
            ws.cell(row=r, column=c+i, value=value)
        r+=1
        #end of init
        for i in range(len(df1)):
            for j in range(len(df2)):
                backup = df1.loc[i].tolist()
                user = df2.loc[j].tolist()
                if backup==user:
                    break
                elif j==len(df2)-1:
                    for k in range(len(backup)):
                        value=backup[k]
                        ws.cell(row=r, column=c+k, value=value)
                    r+=1

    p_bar.grid_forget()
    p_bar.stop()
    SaveBtn.config(state=ACTIVE)

def save(*event):
    wb.save(Save)
    success.grid(row=4, column=0, columnspan=3, padx=10, pady=10)
    success.after(1500, success.grid_forget)

####################################################################################################################################
# Entry Boxes

# Load path for user excel
path_load_user = StringVar()
path_load_user.set("Pfad zur Excel-tabelle auf dem Server")
SearchEntry1 = ttk.Entry(root, textvariable=path_load_user, style="C.TEntry", width=54)
SearchEntry1.grid(row=0, column=1, columnspan=4, padx=10, pady=10)

# Load path for backup Excel
path_load_backup = StringVar()
path_load_backup.set("Pfad zur Backup Excel-Tabelle")
SearchEntry2 = ttk.Entry(root, textvariable=path_load_backup, style="C.TEntry", width=54)
SearchEntry2.grid(row=1, column=1, columnspan=4, padx=10, pady=10)

# Save path for result Excel
path_save = StringVar()
path_save.set("Pfad und Name unter welcher die resultierende Excel-Tabelle gespeichert wird")
SaveEntry = ttk.Entry(root, textvariable=path_save, style="C.TEntry", width=54)
SaveEntry.grid(row=2, column=1, columnspan=4, padx=10, pady=10)

####################################################################################################################################
# Buttons

# Exit button
ExtBtn = ttk.Button(root, text="Exit", command=destroyRoot, style="CExit.TButton")
ExtBtn.grid(row=3, column=0, padx=10, pady=10, sticky=W)
ExtBtn.bind_all("<Escape>", destroyRoot)

# Browse user excel path
SearchBtn1 = ttk.Button(root, text="...", command=loadExcelUser, width=3, style="C.TButton")
SearchBtn1.grid(row=0, column=5, padx=10, pady=10, sticky=E)
SearchBtn1.bind_all("<Return>", loadExcelUser)

# Browse backup excel path
SearchBtn2 = ttk.Button(root, text="...", command=loadExcelBackup, width=3, style="C.TButton")
SearchBtn2.grid(row=1, column=5, padx=10, pady=10, sticky=E)
SearchBtn2.bind_all("<Return>", loadExcelBackup)

# Browse save path and name for result excel
SearchBtn3 = ttk.Button(root, text="...", command=saveResult, width=3, style="C.TButton")
SearchBtn3.grid(row=2, column=5, padx=10, pady=10, sticky=E)
SearchBtn3.bind_all("<Return>", saveResult)

# Execute the comparison
CompareBtn = ttk.Button(root, text="Abgleich", command=ShowWork, style="C.TButton", state=DISABLED)
CompareBtn.grid(row=3, column=5, padx=10, pady=10)
CompareBtn.bind_all("<Return>", compareExcel)

# Save resulting excel table
SaveBtn = ttk.Button(root, text="Speichern", command=save, style="C.TButton", state=DISABLED)
SaveBtn.grid(row=3, column=4, padx=10, pady=10)
SaveBtn.bind_all("<Return>", save)


####################################################################################################################################
# Labels

# user path label
UserPathLabel = ttk.Label(root, text="Pfad zur Server-Excel: ", style="C.TLabel").grid(row=0, column=0, sticky=W, padx=10, pady=10)
# backup path label
BackupPathLabel = ttk.Label(root, text="Pfad zur Backup-Excel: ", style="C.TLabel").grid(row=1, column=0, sticky=W, padx=10, pady=10)
# save path label
UserPathLabel = ttk.Label(root, text="Pfad zur Resultat-Excel: ", style="C.TLabel").grid(row=2, column=0, sticky=W, padx=10, pady=10)
# Progressbar to show background work being done (aka not stuck)
p_bar = ttk.Progressbar(root, orient=HORIZONTAL, length=300, mode="indeterminate")
# Save was successful
success = ttk.Label(root, text="Excel erfolgreich gespeichert", style="C.TLabel")

####################################################################################################################################
# Wrapping
root.mainloop()




































